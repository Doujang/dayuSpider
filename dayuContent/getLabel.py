# coding = utf-8

import openpyxl
import re
import requests
import time
import json

'''
读取Excel文档，抓取每篇文章的命中标签、阅读数以及评论数，并更新入Excel
'''

#读取Excel文件
def readExcel(filename):
    # 打开Excel文件
    wb = openpyxl.load_workbook(filename)
    # 读取第三张表，下标为2，想读取其它表，改变下标即可
    sheet = wb['UC头条']
    # cell_value为第一行第一列单元格的内容，下标为（0,0），想读取其它单元格，改变下标即可
    for i in range(4,413):
        print(i)
        try:
            url_value = sheet.cell(i,5).value
            pattern = re.compile('wm_aid=[^\s]*')
            wm_aid = re.findall(pattern,url_value)
            #print(wm_aid)
            #print(len(wm_aid[0]))
            aid = wm_aid[0][7:39]
            aid = str(aid)
            #print(aid)

            pattern2 = re.compile('wm_id=[^\s]*')
            wm_id = re.findall(pattern2,url_value)
            print(wm_id)
            print(len(wm_id[0]))
            id = wm_id[0][6:38]
            print(id)
        except:
            continue
        cid = loadPage(id,aid)
        page_url = 'http://ff.dayu.com/contents/' + str(cid) + '?biz_id=1002&_fetch_author=1&_fetch_incrs=1'
        items = loadLink(page_url)
        saveExcel(i,items,filename)

def loadPage(mid,aid):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36',
        'Host': 'ff.dayu.com',
        'Connection': 'keep-alive',
        'Cache - Control': 'max - age = 0',
        'Upgrade-Insecure-Requests': '1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        #'Cookie':'cna=ssFmE2lTTh8CAXeDsgjtj/mi; isg=BHZ2ncjXMY0PIcVteei_Cq3Qx6y4P7u9nkIVH-Bf7NnQIxW9SCVZ48LSP7_qi7Lp'
        }
    for i in range(1,50): #翻到第50页即可
        url = 'http://ff.dayu.com/contents/author/' + str(mid) + '?biz_id=1002&_size=8&_page=' + str(i)
        print(url)
        try:
            body = requests.get(url,headers=headers,timeout=5).text
            time.sleep(0.01)
            print(body)
            response = json.loads(body)
            data = response['data']
        except:
            print('something is wrong!!!')
            return

        for i in range(len(data)):
            try:
                oid = data[i]['origin_id']
            except:
                oid = ''
            if oid == aid:
                content_id = data[i]['content_id']
                return content_id


def saveExcel(i,items,filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb['UC头条']
    try:
        label = items['label']
    except:
        label = '[]'
    try:
        read_count = items['read_count']
    except:
        read_count = 0
    try:
        comment_count = items['comment_count']
    except:
        comment_count = 0
    sheet.cell(i,6,label)
    sheet.cell(i,8,read_count)
    sheet.cell(i,9,comment_count)
    wb.save(filename)

def loadLink(source_url):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36',
        'Host': 'ff.dayu.com',
        'Connection': 'keep-alive',
        'Cache - Control': 'max-age=0',
        'Upgrade-Insecure-Requests': '1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        #'Cookie':'cna=ssFmE2lTTh8CAXeDsgjtj/mi; isg=BHZ2ncjXMY0PIcVteei_Cq3Qx6y4P7u9nkIVH-Bf7NnQIxW9SCVZ48LSP7_qi7Lp'
    }
    try:
        body = requests.get(source_url, headers=headers, timeout=5).text
        time.sleep(0.01)
        print(body)
        soup = json.loads(body)
        data = soup['data']
    except:
        print('something is wrong!!!')
        return

    # 阅读数
    try:
        click1 = data['_incrs']['click1']
        click1 = int(click1)
        click2 = data['_incrs']['click2']
        click2 = int(click2)
        click3 = data['_incrs']['click3']
        click3 = int(click3)
        read_count = click1 + click2 + click3
    except:
        read_count = 0

    #评论数
    try:
        comment_count = data['_incrs']['comment']
        comment_count = int(comment_count)
    except:
        comment_count = 0

    #标签数
    try:
        item_id = data['_extra']['xss_item_id']
        label = get_Label(item_id)
    except:
        label = '[]'

    items_list = {
        'read_count': read_count,
        'comment_count': comment_count,
        'label':label
    }

    time.sleep(0.01)
    return items_list

def get_Label(id):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/66.0.3359.139 Safari/537.36',
        'Host': 'm.uczzd.cn',
        'Connection': 'keep-alive',
        'Cache - Control': 'max-age=0',
        'Upgrade-Insecure-Requests': '1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8',
        'Accept-Encoding': 'gzip, deflate',
        'Accept-Language': 'zh-CN,zh;q=0.9',
        # 'Cookie':'cna=ssFmE2lTTh8CAXeDsgjtj/mi; isg=BHZ2ncjXMY0PIcVteei_Cq3Qx6y4P7u9nkIVH-Bf7NnQIxW9SCVZ48LSP7_qi7Lp'
    }
    url = 'http://m.uczzd.cn/iflow/api/v1/article?aid=%5B%22' + id + '%22%5D'
    try:
        body = requests.get(url, headers=headers, timeout=5).text
        time.sleep(0.01)
        print(body)
        soup = json.loads(body)
        label = soup['data']['items'][id]['tags']
        label = str(label)
        print(label)
    except:
        label = '[]'

    return label

if __name__ == "__main__":
    filename = 'dayu.xlsx'
    readExcel(filename)
